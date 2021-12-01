import csv 
from openpyxl import load_workbook
from openpyxl import Workbook

def generate_marksheet_range(sr,er):

    if (sr[0]!=er[0] or sr[1]!=er[1] or sr[2]!=er[2] or sr[3]!=er[3] or sr[4]!=er[4] or sr[5]!=er[5]) :
        return 
    srx = int(str(sr[6]+sr[7]))
    erx = int(str(er[6]+er[7]))

    if srx > erx :
        return 

    print(1)
    names = {}                                                                   #dictionary with roll no as key and name as value
    subject = {}                                                                 #dictionary with subject as key and subject attributes as value in form of list
    spi = {}                                                                     #dictionary with rollnumber+sem as key and spi as value
    cpi = {}                                                                     #dictionary with rollnumber+sem as key and cpi as value
    credits_earned = {}                                                          #dictionary with rollnumber+sem as key and credits earned as value
    tot_credits = {}                                                             #dictionary with rollnumber+sem as key and total credits as value
    marks = {}                                                                   #dictionary with grade in alphabets as key and grade in numbers as value
    marks["AA"] = 10
    marks["AB"] = 9
    marks["BB"] = 8
    marks[" BB"] = 8
    marks["BC"] = 7
    marks["CC"] = 6
    marks["CD"] = 5
    marks["DD"] = 4
    marks["DD*"] = 4
    marks["F"] = 0
    marks["F*"] = 0
    marks["I"] = 0
    with open('./sample_input/names-roll.csv','r') as file:                                         #File opening
        f_in = csv.reader(file)                                                      #File Reading
        
        for elements in f_in :                                                       #loop for each row in given file
            names[elements[0]] = elements[1]

    with open('./sample_input/subjects_master.csv','r') as file:                                     #File opening
        f_in = csv.reader(file)                                                       #File Reading
        
        for elements in f_in :                                                       #loop for each row in given file
            subject[elements[0]] = [elements[1],elements[2],elements[3]]

    roll_dict = {}                                                                   #dictionary for maintaining Key : roll numbers
    with open('./sample_input/grades.csv','r') as file:                                             #File opening
        f_in = csv.reader(file)                                                      #File Reading
        cnt = 0                                                                      #Represents current row in file
        
        for elements in f_in :  

            rollsem = str(elements[0]) + str(elements[1]) 
                                                                                     #loop for each row in given file
            if cnt == 0 :
                cnt += 1
                continue

            #print (elements[0],sr)

            # if cnt == 3 :
            #     break

            if (elements[0][0] != sr[0] or elements[0][1] != sr[1] or elements[0][2] != sr[2] or elements[0][3] != sr[3] or elements[0][4] != sr[4] or elements[0][5] != sr[5]) :
                cnt += 1
                continue
            #print (cnt)
            
            st = int (str(sr[6] + sr[7]))
            se = int (str(er[6] + er[7]))
            crr = int (str(elements[0][6] + elements[0][7]))

            if (crr < st or crr > se) :
                cnt += 1
                continue

            #print(sr,er,crr)

            cur_roll = elements[0]                                                   #represents current roll number     
            
            if int(elements[1]) > 8 :
                continue

            if cur_roll not in roll_dict :
                roll_dict[cur_roll] = 1                                             #if the Roll no appears 1st time
                
                #initiallising values
                for x in range (1,9) :
                    rsem=cur_roll + str(x)
                    spi[rsem] = 0
                    credits_earned[rsem] = 0
                    tot_credits[rsem] = 0
                
                cpi[rollsem] = 0
                wb = Workbook()                                                     #opening the file
                sheet = wb.active                                                   
                sheet.title = "Overall"                                             #creating different sheets 
                for x in range (1,9) :                                              #creating different sheets
                    sheetname = "Sem"+str(x)
                    wb.create_sheet(index=x,title=sheetname)
                    sheet=wb[sheetname]
                    #writting first row of each sheet
                    sheet.cell(row = 1, column = 1).value = 'Sl No.'                
                    sheet.cell(row = 1, column = 2).value = 'Subject No.'
                    sheet.cell(row = 1, column = 3).value = 'Subject Name'
                    sheet.cell(row = 1, column = 4).value = 'L-T-P'
                    sheet.cell(row = 1, column = 5).value = 'Credit'
                    sheet.cell(row = 1, column = 6).value = 'Subject Type'
                    sheet.cell(row = 1, column = 7).value = 'Grade'
                
                sheet = wb["Overall"]
                #writting first column of "Overall" sheet
                sheet.cell(row = 1, column = 1).value = 'Rollno'
                sheet.cell(row = 1, column = 2).value = cur_roll
                sheet.cell(row = 2, column = 1).value = 'Name of Student'
                sheet.cell(row = 2, column = 2).value = names[cur_roll]
                sheet.cell(row = 3, column = 1).value = 'Discipline'
                sheet.cell(row = 3, column = 2).value = cur_roll[4]+cur_roll[5]
                sheet.cell(row = 4, column = 1).value = 'Semester No.'
                for y in range (1,9) :
                    sheet.cell(row = 4, column = 1+y).value = y
                sheet.cell(row = 5, column = 1).value = 'Semester wise Credit Taken'
                sheet.cell(row = 6, column = 1).value = 'SPI'
                sheet.cell(row = 7, column = 1).value = 'Total Credits Taken'
                sheet.cell(row = 8, column = 1).value = 'CPI'
                
                wb.save("output/" + cur_roll + ".xlsx")                             #saving the file in the given path

            
            path = "output/" + cur_roll + ".xlsx"                                   #assigning path for writing
            wb = load_workbook(path)                                                #opening the file of the given path
            sheetname = "Sem"+str(elements[1])
            sheet=wb[sheetname]
            row_no = sheet.max_row                                                  #finding maximum row available
            row_no += 1
            sheet.cell(row = row_no, column = 1).value = row_no-1                   
            sheet.cell(row = row_no, column = 2).value = elements[2]
            sheet.cell(row = row_no, column = 3).value = subject[elements[2]][0]
            sheet.cell(row = row_no, column = 4).value = subject[elements[2]][1]
            sheet.cell(row = row_no, column = 5).value = elements[3]
            sheet.cell(row = row_no, column = 6).value = elements[5]
            sheet.cell(row = row_no, column = 7).value = elements[4]
            credits_earned[rollsem] += marks[elements[4]]*(int(elements[3]))        #updating score
            tot_credits[rollsem] += int(elements[3])                                #updating total credits till now
            wb.save(path)
            cnt += 1
        
        cntt = 0

        for x, y in roll_dict.items():
            
            #print (x)
            # if cntt == 3 :
            #     break
            cntt += 1
            
            path = "output/" + x + ".xlsx"                                          #assigning path for writing
            wb = load_workbook(path)                                                #opening the file of the given path
            sheetname = "Overall"
            sheet=wb[sheetname]
            xx = 0
            yy = 0
            for z in range(1,9) :
                rollsem = str(x)+str(z)
                if int(tot_credits[rollsem]) == 0 :
                    sheet.cell(row = 5, column = z+1).value = 0
                    sheet.cell(row = 6, column = z+1).value = 0       
                else : 
                    xx += credits_earned[rollsem]
                    yy += tot_credits[rollsem]
                    sheet.cell(row = 5, column = z+1).value = tot_credits[rollsem]
                    sheet.cell(row = 6, column = z+1).value = round(credits_earned[rollsem]/tot_credits[rollsem],2)       #calculating spi
                if yy == 0 :
                    sheet.cell(row = 7, column = z+1).value = 0
                    sheet.cell(row = 8, column = z+1).value = 0
                else :
                    sheet.cell(row = 7, column = z+1).value = yy
                    sheet.cell(row = 8, column = z+1).value = round(xx/yy,2)                                              #calculating cpi
            wb.save(path)                                                                   #saving file
    return
