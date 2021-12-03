from create_table_fpdf2 import PDF
import pandas as pd
import os
import openpyxl

def generate() :
    for file in os.listdir("output"):
        print (file)
        path = os.path.join("output" , file)
        wb = openpyxl.load_workbook(path,'rb')

        pdf = PDF('L', 'mm', 'A3')
        pdf.add_page()
        pdf.set_font("Times", size=10)

        sheet = wb["Overall"]
        roll = sheet.cell(1,2).value
        st1 = "Roll Number " + sheet.cell(1,2).value + "     Name: " +  sheet.cell(2,2).value + "     Year of Admission: 2019 "

        sst = "Bachelor of Technology"
        if roll[2] == '1' :
            sst = "Master of Technology"

        dict= {'CS': 'Computer Science and Engineering', 'EE': 'Electrical Engineering', 'CB': 'Chemical Engineering', 'CE':'Civil Engineering','MM':'Mechanical Engineering','MME':'Metallurgy and Materials Engineering'}

        ts = roll[4] + roll[5]

        sst1 = ts
        if ts in dict :
            sst1 = dict [ts]

        st2 = "Programme: " +  sst + "    Course: "  +   sst1



        pdf.header(st1,st2)

        for semi in range (1,9) :

            sem = "Sem" + str(semi)
            sheet = wb[sem]
            tot_r = sheet.max_row



            clear = 0

            data = []

            for i in range (1,tot_r) :
                row = []
                if i != 1 :
                    clear += int(sheet.cell(i,5).value) * int((str(sheet.cell(i,j).value) != 'F'))
                for j in range (2,8) :
                    if j == 6:
                        continue 
                    row.append(str(sheet.cell(i,j).value))
                data.append(row)
            ff = 0
            
            if semi % 4 == 1 :
                ff= +9.999999999999998
            sheet = wb["Overall"]
            sst = "Credits taken: " + str(sheet.cell(7,semi+1).value) + "     Credits Cleared:  " + str(clear) + "     SPI: " + str(sheet.cell(6,semi+1).value) + "     CPI: " + str(sheet.cell(8,semi+1).value)
            if (len(data) == 0) :
                continue
            pdf.create_table(table_data = data,title="Semester" + str(semi), x_start = ((semi-1)%4)*100 + ff,y_start = 50 + ((semi-1)//4) *70, f = semi%2,cell_width='uneven',st = sst)

        op = os.path.join("transcriptsIITP" , roll + ".pdf" )

        pdf.output(op)