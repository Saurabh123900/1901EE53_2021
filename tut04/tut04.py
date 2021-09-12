import csv 
from openpyxl import load_workbook
from openpyxl import Workbook


def output_by_subject():
    subject_dict = {}                                                                #dictionary for maintaining Key : subject, value : Row number in .xlsx file
    with open('regtable_old.csv','r') as file:                                       #File opening
        f_in = csv.reader(file)                                                      #File Reading
        cnt = 0                                                                      #Represents current row in file
        
        for elements in f_in :                                                       #loop for each row in given file
            if cnt == 0 :
                cnt += 1
                continue
            cur_subject = elements[3]                                                        
            
            if cur_subject not in subject_dict :                                    #if the subject appears 1st time
                wb = Workbook()                                                     #opening the file
                sheet = wb.active                                                   
                sheet.cell(row = 1, column = 1).value = 'rollno'
                sheet.cell(row = 1, column = 2).value = 'register_sem'
                sheet.cell(row = 1, column = 3).value = 'subno'
                sheet.cell(row = 1, column = 4).value = 'sub_type'
                subject_dict[cur_subject] = 2
                wb.save("output_by_subject/" + cur_subject + ".xlsx")               #saving the file in the given path

            
            path = "output_by_subject/" + cur_subject + ".xlsx"                      #assigning path for writing
            wb = load_workbook(path)                                                 #opening the file of the given path
            sheet = wb.active
            cur_row = subject_dict[cur_subject]
            
            sheet.cell(row = cur_row, column = 1).value = elements[0]              #writting the reuired imformation
            sheet.cell(row = cur_row, column = 2).value = elements[1]              #writting the reuired imformation
            sheet.cell(row = cur_row, column = 3).value = elements[3]              #writting the reuired imformation
            sheet.cell(row = cur_row, column = 4).value = elements[8]              #writting the reuired imformation
            subject_dict[cur_subject] = cur_row + 1
            wb.save(path)
    return

def output_individual_roll():
    roll_dict = {}                                                                   #dictionary for maintaining Key : roll numbers, value : Row number in .xlsx file
    with open('regtable_old.csv','r') as file:                                       #File opening
        f_in = csv.reader(file)                                                      #File Reading
        cnt = 0                                                                      #Represents current row in file
        
        for elements in f_in :                                                       #loop for each row in given file
            if cnt == 0 :
                cnt += 1
                continue
            cur_roll = elements[0]                                                        
            
            if cur_roll not in roll_dict :                                          #if the Roll no appears 1st time
                wb = Workbook()                                                     #opening the file
                sheet = wb.active                                                   
                sheet.cell(row = 1, column = 1).value = 'rollno'
                sheet.cell(row = 1, column = 2).value = 'register_sem'
                sheet.cell(row = 1, column = 3).value = 'subno'
                sheet.cell(row = 1, column = 4).value = 'sub_type'
                roll_dict[cur_roll] = 2
                wb.save("output_individual_roll/" + cur_roll + ".xlsx")                  #saving the file in the given path

            
            path = "output_individual_roll/" + cur_roll + ".xlsx"                        #assigning path for writing
            wb = load_workbook(path)                                                #opening the file of the given path
            sheet = wb.active
            cur_row = roll_dict[cur_roll]
            
            sheet.cell(row = cur_row, column = 1).value = elements[0]              #writting the reuired imformation
            sheet.cell(row = cur_row, column = 2).value = elements[1]              #writting the reuired imformation
            sheet.cell(row = cur_row, column = 3).value = elements[3]              #writting the reuired imformation
            sheet.cell(row = cur_row, column = 4).value = elements[8]              #writting the reuired imformation
            roll_dict[cur_roll] = cur_row + 1
            wb.save(path)
    return
    
output_by_subject()                                                              #Function Call
output_individual_roll()                                                         #Function Call