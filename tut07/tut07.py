import csv
from re import sub 
from openpyxl import load_workbook
from openpyxl import Workbook


def feedback_not_submitted():

	
	ltp_mapping_feedback_type = {1: 'lecture', 2: 'tutorial', 3:'practical'}
	output_file_name = "course_feedback_remaining.xlsx" 
	
	subject = {}					#(subject name) => all info of subject (ltc)	   
	rollsub = {}					#roll + subject + degree(ltc)  =>  1
	rollsuba = {}					#roll  =>  list of all subjects
	stud = {}						#roll  =>  student info full
	rsub = {}						#roll+sub => reg sems  
	visrsub = {}					#roll + sub => bool

	with open("course_master_dont_open_in_excel.csv",'r') as file:
		f = csv.reader(file)                                                                                            # open the file for reading                                                                                                       # dictionary to store key value pairs of current subject and the row number at which we have to write in the excel file with name current_subject
		cnt  = 0                                                                                                        # checks if we are at row 0 in our main file
		for row in f: 
			if cnt == 0:                                                                                                # if we are currently at row 0, it means we are at headers, we continue
				cnt += 1                                                                                                 # we set the cnt  to be 1, which indicates that we are not at row 0 anymore
				continue
			a = row[2].split('-')
			subject[row[0]] = a
			cnt+=1

	with open("course_feedback_submitted_by_students.csv",'r') as file:
		f = csv.reader(file)                                                                                            # open the file for reading                                                                                                       # dictionary to store key value pairs of current subject and the row number at which we have to write in the excel file with name current_subject
		cnt  = 0                                                                                                        # checks if we are at row 0 in our main file
		for row in f: 
			if cnt == 0:                                                                                                # if we are currently at row 0, it means we are at headers, we continue
				cnt += 1                                                                                                 # we set the cnt  to be 1, which indicates that we are not at row 0 anymore
				continue
			s = row[3] + row[4] + row[5]
			rollsub[s] = 1
			cnt+=1

	with open("course_registered_by_all_students.csv",'r') as file:
		f = csv.reader(file)                                                                                            # open the file for reading                                                                                                       # dictionary to store key value pairs of current subject and the row number at which we have to write in the excel file with name current_subject
		cnt  = 0                                                                                                        # checks if we are at row 0 in our main file
		for row in f: 
			if cnt == 0:                                                                                                # if we are currently at row 0, it means we are at headers, we continue
				cnt = 1                                                                                                 # we set the cnt  to be 1, which indicates that we are not at row 0 anymore
				continue
			if row[0] in rollsuba :
				rollsuba[row[0]].append(row[3])
			else :
				rollsuba[row[0]] = []
				rollsuba[row[0]].append(row[3])
			s = row[0] + row[3]
			rsub[s] = [row[1],row[2]]
			cnt+=1

	with open("studentinfo.csv",'r') as file:
		f = csv.reader(file)                                                                                            # open the file for reading                                                                                                       # dictionary to store key value pairs of current subject and the row number at which we have to write in the excel file with name current_subject
		cnt  = 0                                                                                                        # checks if we are at row 0 in our main file
		for row in f: 
			if cnt == 0:                                                                                                # if we are currently at row 0, it means we are at headers, we continue
				cnt = 1                                                                                                 # we set the cnt  to be 1, which indicates that we are not at row 0 anymore
				continue
			stud[row[1]] = row
			cnt+=1
	
	output_file_name = "course_feedback_remaining.xlsx"
	wb = load_workbook(output_file_name) 
	sheet = wb.active
	cur = 2                                                                                   
	sheet['A1'] = 'rollno'
	sheet['B1'] = 'register_sem'
	sheet['C1'] = 'schedule_sem'
	sheet['D1'] = 'subno'
	sheet['E1'] = 'Name'
	sheet['F1'] = 'email'
	sheet['G1'] = 'aemail'
	sheet['H1'] = 'contact'
	wb.save(output_file_name)

	for i in rollsuba :
		for j in rollsuba[i] : 
			ss = i + j
			check = 1
			for k in range (3) :
				if subject[j][k] == "0" :
				 	continue
				s = i + j + str(k+1)
				if s in rollsub :
					continue
				check = 0
				break
			if ss in visrsub :
				continue
			visrsub[ss] = 1
			if check == 0 :
				sheet.cell(row = cur, column = 1).value = i
				sheet.cell(row = cur, column = 2).value = rsub[ss][0]
				sheet.cell(row = cur, column = 3).value = rsub[ss][1]
				sheet.cell(row = cur, column = 4).value = j
				if i in stud :
					sheet.cell(row = cur, column = 5).value = stud[i][0]
					sheet.cell(row = cur, column = 6).value = stud[i][8]
					sheet.cell(row = cur, column = 7).value = stud[i][9]
					sheet.cell(row = cur, column = 8).value = stud[i][10]
				else :
					sheet.cell(row = cur, column = 5).value = "NA"
					sheet.cell(row = cur, column = 6).value = "NA"
					sheet.cell(row = cur, column = 7).value = "NA"
					sheet.cell(row = cur, column = 8).value = "NA"
				cur += 1
	wb.save(output_file_name)





feedback_not_submitted()
